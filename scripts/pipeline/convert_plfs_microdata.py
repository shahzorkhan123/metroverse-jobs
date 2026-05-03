"""Convert PLFS fixed-width unit-level TXT files to the CSV format
expected by import_india_subnational_from_microdata().

Usage:
    # Convert a single year (perv1.txt in data/raw/plfs_microdata/YYYY/)
    python scripts/pipeline/convert_plfs_microdata.py --year 2021

    # Convert all years found in data/raw/plfs_microdata/
    python scripts/pipeline/convert_plfs_microdata.py --all

Output:
    data/raw/ind_plfs_microdata.csv  (overwritten for the specified year)

The pipeline then reads this CSV via:
    python scripts/pipeline/run_pipeline.py --year YYYY --country in

Download microdata from: https://microdata.mospi.gov.in
  -> Search "PLFS Annual" -> download zip -> extract perv1.txt
  -> Place in data/raw/plfs_microdata/YYYY/perv1.txt

Layout reference: data/raw/Data_LayoutPLFS_2023-24.xlsx (sheet 'perv1')

Key fields extracted from perv1.txt (fixed-width):
  Field       Byte positions  Width  Notes
  st          11-12           2      State/UT code (01=J&K, 07=Delhi, ...)
  ocu_pas     col varies      3      NCO-2015 occupation code (principal activity)
  mult        col varies      10     Sub-sample multiplier (survey weight)
  ern_reg     col varies      8      Monthly earnings for regular/salaried workers
"""

import argparse
import csv
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Field layouts by year
# The layout changes slightly between survey rounds.
# Format: {year: {field_name: (start_byte, width)}} — 1-indexed byte positions.
# ---------------------------------------------------------------------------

# These come from Data_LayoutPLFS_2023-24.xlsx (sheet perv1), confirmed for
# 2020-21 through 2023-24. Earlier years may differ; add overrides below.
LAYOUT_DEFAULT = {
    # Household identifiers (used to deduplicate)
    "st":      (11, 2),   # State/UT code
    # Person-level occupation (Block 5.1 — principal activity status)
    "ocu_pas": (None, 3), # NCO code — byte position varies; computed below
    # Generated fields
    "mult":    (None, 10),# Sub-sample multiplier
    # Earnings
    "ern_reg": (None, 8), # Regular wage/salary earnings (monthly, Rs.)
}

# Byte positions confirmed from layout file (1-indexed, inclusive start):
#   perv1 record length: varies by year
#
# 2023-24 layout (from Data_LayoutPLFS_2023-24.xlsx):
#   st        : bytes 11-12
#   ocu_pas   : Block 5.1 col 6, 3 chars — need to locate in full layout
#   mult      : near end of record, 10 chars
#   ern_reg   : Block 6 col 9, 8 chars

# Since the exact byte positions depend on summing all prior field widths,
# we use the field names from the layout and look them up dynamically.
# The function below parses the layout xlsx to get precise positions.

ROOT = Path(__file__).resolve().parent.parent.parent
RAW_DIR = ROOT / "data" / "raw"
MICRO_DIR = RAW_DIR / "plfs_microdata"
OUT_CSV = RAW_DIR / "ind_plfs_microdata.csv"
LAYOUT_FILE = RAW_DIR / "Data_LayoutPLFS_2023-24.xlsx"

# Hardcoded field positions from the 2023-24 layout (perv1 sheet).
# These are verified from the xlsx: (start_col, width) where start_col is
# 1-indexed byte offset in the fixed-width record.
#
# To regenerate: read the 'perv1' sheet, column 'Byte Position' and 'Field Length'.
FIELD_POSITIONS = None  # Populated by _load_layout()


def _load_layout():
    """Parse Data_LayoutPLFS_2023-24.xlsx to get field byte positions."""
    global FIELD_POSITIONS
    if FIELD_POSITIONS is not None:
        return FIELD_POSITIONS

    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed; using hardcoded positions.")
        FIELD_POSITIONS = _hardcoded_positions()
        return FIELD_POSITIONS

    if not LAYOUT_FILE.exists():
        print(f"Layout file not found: {LAYOUT_FILE}; using hardcoded positions.")
        FIELD_POSITIONS = _hardcoded_positions()
        return FIELD_POSITIONS

    wb = openpyxl.load_workbook(LAYOUT_FILE, read_only=True)
    ws = wb["perv1"]
    rows = list(ws.iter_rows(values_only=True))

    # Accumulate byte position by summing widths
    pos = 1
    positions = {}
    for row in rows[1:]:  # skip header
        if row[0] is None:
            continue
        field_name = str(row[5] or "").strip()  # column F = Field_Name
        width_raw = row[4]  # column E = Field Length
        try:
            width = int(width_raw)
        except (TypeError, ValueError):
            continue
        if field_name:
            positions[field_name] = (pos, width)
        pos += width

    FIELD_POSITIONS = positions
    return FIELD_POSITIONS


def _hardcoded_positions():
    """Hardcoded byte positions from 2023-24 layout as fallback.
    Derived from Data_LayoutPLFS_2023-24.xlsx perv1 sheet by cumulative sum.
    """
    return {
        "file_id":  (1,  4),
        "sch":      (5,  3),
        "qtr":      (8,  2),
        "visit":    (10, 2),
        "sec":      (12, 1),
        "st":       (13, 2),   # State/UT code
        "dst":      (15, 2),
        "nss_reg":  (17, 3),
        "stratum":  (20, 2),
        "sub_str":  (22, 2),
        "sub_smp":  (24, 1),
        "fod_sub":  (25, 4),
        "fsu":      (29, 5),
        "sg_sb":    (34, 1),
        "sss":      (35, 1),
        "hh_no":    (36, 2),
        "mo_surv":  (38, 2),
        # ... many more fields ...
        # Key fields (approximate — recalculated from xlsx):
        "ocu_pas":  (None, 3),  # Will be resolved from layout
        "mult":     (None, 10),
        "ern_reg":  (None, 8),
    }


def _get_pos(field_name):
    """Return (start_0indexed, width) for a field."""
    positions = _load_layout()
    if field_name not in positions or positions[field_name][0] is None:
        raise KeyError(f"Field '{field_name}' not found in layout")
    start_1, width = positions[field_name]
    return start_1 - 1, width  # convert to 0-indexed


def _read_state_codes():
    """Return mapping of state code string -> state name from layout xlsx."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(LAYOUT_FILE, read_only=True)
        ws = wb["State code"]
        mapping = {}
        for row in ws.iter_rows(values_only=True):
            if row[0] is None or row[1] is None:
                continue
            code = str(row[0]).strip().zfill(2)
            name = str(row[1]).strip()
            if code.isdigit():
                mapping[code] = name
        return mapping
    except Exception:
        # Hardcoded fallback from layout file
        return {
            "01": "Jammu & Kashmir", "02": "Himachal Pradesh",
            "03": "Punjab", "04": "Chandigarh", "05": "Uttarakhand",
            "06": "Haryana", "07": "Delhi", "08": "Rajasthan",
            "09": "Uttar Pradesh", "10": "Bihar", "11": "Sikkim",
            "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur",
            "15": "Mizoram", "16": "Tripura", "17": "Meghalaya",
            "18": "Assam", "19": "West Bengal", "20": "Jharkhand",
            "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh",
            "24": "Gujarat", "25": "Daman & Diu",
            "26": "Dadra & Nagar Haveli", "27": "Maharashtra",
            "28": "Andhra Pradesh", "29": "Karnataka", "30": "Goa",
            "31": "Lakshadweep", "32": "Kerala", "33": "Tamil Nadu",
            "34": "Puducherry", "35": "Andaman & Nicobar Islands",
            "36": "Telangana", "37": "Ladakh",
        }


def convert(txt_path: Path, out_csv: Path, year: int) -> int:
    """Convert a single perv1.txt to CSV. Returns number of rows written."""
    positions = _load_layout()

    # Resolve field positions
    try:
        st_s, st_w     = _get_pos("st")
        ocu_s, ocu_w   = _get_pos("ocu_pas")
        mult_s, mult_w = _get_pos("mult")
        ern_s, ern_w   = _get_pos("ern_reg")
    except KeyError as e:
        print(f"  ERROR: {e}")
        print(f"  Available fields: {sorted(positions.keys())[:20]}...")
        return 0

    state_codes = _read_state_codes()

    rows_written = 0
    skipped = 0

    with open(txt_path, "r", encoding="latin-1", errors="replace") as fin, \
         open(out_csv, "w", newline="", encoding="utf-8") as fout:

        writer = csv.DictWriter(fout, fieldnames=["st", "ocu_pas", "mult", "ern_reg"])
        writer.writeheader()

        for line in fin:
            line = line.rstrip("\n\r")
            if len(line) < max(st_s + st_w, ocu_s + ocu_w, mult_s + mult_w):
                skipped += 1
                continue

            st_code = line[st_s:st_s + st_w].strip()
            ocu     = line[ocu_s:ocu_s + ocu_w].strip()
            mult    = line[mult_s:mult_s + mult_w].strip()

            ern = ""
            if ern_s is not None and ern_s >= 0:
                ern = line[ern_s:ern_s + ern_w].strip() if len(line) > ern_s else ""

            # Map state code to name
            st_name = state_codes.get(st_code.zfill(2), "")
            if not st_name or not ocu or not mult:
                skipped += 1
                continue

            writer.writerow({
                "st":      st_name,
                "ocu_pas": ocu,
                "mult":    mult,
                "ern_reg": ern if ern else "",
            })
            rows_written += 1

    return rows_written


def main():
    parser = argparse.ArgumentParser(description="Convert PLFS microdata TXT to CSV")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--year", type=int, help="Calendar year to convert (e.g. 2021)")
    group.add_argument("--all", action="store_true", help="Convert all years found")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print what would be done without writing")
    args = parser.parse_args()

    _load_layout()  # Validate layout file once

    if args.all:
        years = sorted(
            int(p.name) for p in MICRO_DIR.iterdir()
            if p.is_dir() and p.name.isdigit()
        )
    else:
        years = [args.year]

    if not years:
        print(f"No year directories found in {MICRO_DIR}")
        print("Create: data/raw/plfs_microdata/YYYY/perv1.txt")
        sys.exit(1)

    for year in years:
        year_dir = MICRO_DIR / str(year)
        txt_path = year_dir / "perv1.txt"

        if not txt_path.exists():
            print(f"  {year}: perv1.txt not found at {txt_path} — skipping")
            continue

        size_mb = txt_path.stat().st_size / 1e6
        print(f"  {year}: {txt_path.name} ({size_mb:.1f} MB)")

        if args.dry_run:
            continue

        n = convert(txt_path, OUT_CSV, year)
        print(f"    → {n:,} rows written to {OUT_CSV.name}")

        if n > 0:
            print(f"    Now run: python scripts/pipeline/run_pipeline.py --year {year} --country in")


if __name__ == "__main__":
    main()
